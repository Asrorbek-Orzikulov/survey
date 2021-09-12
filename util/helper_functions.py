import tkinter as tk
import util
import os
import xlsxwriter
from tkinter import messagebox
from glob import glob
import openpyxl


def is_digit(input_string, action_type):
    if action_type == '1':  # insertion
        if not input_string.isdigit():
            return False
    return True


def is_cyrillic(input_string, action_type):
    if action_type == '1':
        if input_string not in util.ALPHABET:
            return False
    return True


def create_messagebox(text, is_error=True):
    if is_error:
        messagebox.showerror(title="Saving Request", message=text)
    else:
        messagebox.showinfo(title="Saving Request", message=text)


def is_properly_filled(input_widgets):
    interviewer_id, mahalla_id, survey_id = input_widgets[0][0]
    if any(
        (len(interviewer_id.get()) != 4,
         len(mahalla_id.get()) != 2,
         len(survey_id.get()) != 2)
    ):
        create_messagebox("ID рақамларни текширинг.")
        # util.log('error', "ID рақамларни текширинг.")
        return False

    for question_num, (element, state) in enumerate(input_widgets[:-1]):

        # early stopping conditions
        if question_num == 2:
            if element.get() == 2:
                return True
        elif question_num == 4:
            if element.get() == 3:
                return True
        elif question_num == 6:
            age = element.get()
            if age != "":
                if int(age) < 18:
                    return True

        # skipping questions based on previous answers
        elif question_num in [36, 39]:
            previous_question = input_widgets[question_num-1][0]
            if previous_question.get() == 5:
                continue
        elif question_num == 45:
            previous_question = input_widgets[question_num-1][0][0]
            if previous_question.get() not in range(1, 6):
                continue
        elif question_num == 57:
            previous_question = input_widgets[question_num-1][0]
            if previous_question.get() not in range(1, 4):
                continue

        # input verification based on the question type
        if isinstance(element, tk.Entry):
            if element.get() == "":
                create_messagebox(f"{question_num} саволда жавобни киритинг.")
                # util.log('error', f"{question_num} саволда жавобни киритинг.")
                return False
        elif isinstance(element, tk.IntVar):
            if element.get() == 0:
                create_messagebox(f"{question_num} саволда жавобни танланг.")
                # util.log('error', f"{question_num} саволда жавобни танланг.")
                return False
        elif isinstance(element, list):
            answers = []
            if question_num in util.CHECKBOXES:
                for widget in element[:-1]:
                    not_checked = widget.get() == 0
                    answers.append(not_checked)
                is_empty = element[-1].get() == ""
                answers.append(is_empty)
                if all(answers):
                    create_messagebox(f"{question_num} саволда камида битта жавобни танланг.")
                    # util.log('error', f"{question_num} саволда камида битта жавобни танланг.")
                    return False
            elif question_num in util.ENTRY_PLUS_CHECKBOXES:
                checkbox_checked = element[-1].get() == 1
                if checkbox_checked:
                    empty_entries = []
                    for entry in element[:-1]:
                        is_empty = entry.get() == ""
                        empty_entries.append(is_empty)
                    if not all(empty_entries):
                        create_messagebox(f"{question_num} саволда очиқ жавоблар бўш қолсин.")
                        # util.log('error', f"{question_num} саволда очиқ жавоблар бўш қолсин.")
                        return False
                else:
                    entry = element[0]
                    if entry.get() == "":
                        create_messagebox(f"{question_num} саволда камида битта жавобни киритинг.")
                        # util.log('error', f"{question_num} саволда камида битта жавобни киритинг.")
                        return False
            elif question_num in util.RADIO_PLUS_OTHERS:
                radio_button, entry_other = element
                radio_not_selected = radio_button.get() == 0
                entry_empty = entry_other.get() == ""
                if radio_not_selected and entry_empty:
                    create_messagebox(f"{question_num} саволда битта жавобни танланг.")
                    # util.log('error', f"{question_num} саволда битта жавобни танланг.")
                    return False
                if (not radio_not_selected) and (not entry_empty):
                    create_messagebox(f"{question_num} саволда очиқ жавоб бўш қолсин.")
                    # util.log('error', f"{question_num} саволда очиқ жавоб бўш қолсин.")
                    return False
    return True


def clear_data(input_widgets):
    for widgets, state in input_widgets:
        if isinstance(widgets, list):
            for widget in widgets:
                if isinstance(widget, tk.Entry) and state == "hidden":
                    widget.delete(0, tk.END)
                    widget.insert(0, "")
                    widget.configure(state=tk.DISABLED)
                elif isinstance(widget, tk.Entry):
                    widget.delete(0, tk.END)
                    widget.insert(0, "")
                elif isinstance(widget, tk.IntVar) or isinstance(widget, tk.Checkbutton):
                    widget.set(0)
                else:
                    raise ValueError(str(widget.get()) + str(type(widget)))
        elif isinstance(widgets, tk.IntVar):
            widgets.set(0)
        elif isinstance(widgets, tk.Entry):
            widgets.delete(0, tk.END)
            widgets.insert(0, "")
        else:
            raise ValueError(str(widgets.get()) + str(type(widgets)))


def get_info(input_widgets, column_names, input_info):
    """
    Get input information from `input_widgets` and generate
    column names for an Excel file.

    """
    for question_num, (widgets, state) in enumerate(input_widgets):
        if isinstance(widgets, list):
            if len(widgets) > 1:
                for option_num, widget in enumerate(widgets):
                    input_info.append(widget.get())
                    column_name = "Q_" + str(question_num) + "_" + str(option_num + 1)
                    column_names.append(column_name)
    
            elif len(widgets) == 1:
                column_name = "Q_" + str(question_num)
                column_names.append(column_name)
        else:
            input_info.append(widgets.get())
            column_name = "Q_" + str(question_num)
            column_names.append(column_name)

    
def save_data(input_widgets):
    """
    Save data in `input_widgets` to an Excel file.

    """
    if not is_properly_filled(input_widgets):
        return

    column_names = []
    input_info = []
    get_info(input_widgets, column_names, input_info)
    if not os.path.isdir("Database"):
        os.mkdir("Database")
    try:
        os.chdir("Database")
        start_str = ""
        for entry in input_widgets[0][0]:
            start_str += entry.get() + "-"

        count = 0
        file_exists = True
        while file_exists:
            name = start_str + f"{count}.xlsx"
            if os.path.isfile(name):
                count += 1
            else:
                file_exists = False

        with xlsxwriter.Workbook(f"{name}.xlsx") as workbook:
            worksheet = workbook.add_worksheet()
            worksheet.write_row(0, 0, column_names)
            worksheet.write_row(1, 0, input_info)

        create_messagebox(f"Жавобларингиз {name}.xlsx тарзида сақланди.", False)
        # util.log('success', f"Жавобларингиз {name}.xlsx тарзида сақланди.")
        os.chdir("..")
        clear_data(input_widgets)
    except Exception as error:
        create_messagebox(str(error))
        # util.log('error', str(error))


def merge_files():
    if not os.path.isdir("Database"):
        create_messagebox("Database папкаси мавжуд эмас.")
        return

    try:
        os.chdir("Database")
        extension = "xlsx"
        all_filenames = [i for i in glob('*.{}'.format(extension))]
        if not all_filenames:
            create_messagebox("Database папкаси бўш.")
            os.chdir("..")
            return
        elif os.path.exists("merged_file.xlsx"):
            create_messagebox("merged_file.xlsx папкада мавжуд.")
            os.chdir("..")
            return

        with xlsxwriter.Workbook("merged_file.xlsx") as workbook_full:
            worksheet = workbook_full.add_worksheet()
            current_row = 0
            for idx, file in enumerate(all_filenames):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                max_col = sheet.max_column
                if idx == 0:
                    start_row = 1
                else:
                    start_row = 2
        
                for row in sheet.iter_rows(min_row=start_row, max_col=max_col, values_only=True):
                    worksheet.write_row(current_row, 0, row)
                    current_row += 1
        
                workbook.close()

        create_messagebox("Барча жавоблар бирлаштирилди.", False)
        os.chdir("..")

    except Exception as error:
        create_messagebox(str(error))


def entry_clicked(radio_var, entry):
    radio_var.set(0)
    entry.configure(state=tk.NORMAL)


def radio_button_question(frame_name, label_row, options, text, has_other=False):
    label = tk.Label(frame_name, text=text)
    label.grid(row=label_row, column=0, columnspan=3, padx=10, pady=30)
    label.config(font=("Arial", 12))
    variable = tk.IntVar(frame_name)
    for idx, option in enumerate(options):
        button = tk.Radiobutton(frame_name, text=option, value=idx+1, variable=variable)
        button.configure(font=("Arial", 12))
        button.grid(row=label_row+idx+1, column=1, padx=10, pady=5, sticky="w")

    if has_other:
        offset = len(options) + 1
        entry = tk.Entry(frame_name, width=100, validate="key")
        entry.configure(validatecommand=(entry.register(is_cyrillic),'%S','%d'))
        entry.grid(row=label_row+offset+1, column=0, columnspan=3, padx=10, pady=15)
        entry.configure(state=tk.DISABLED, font=("Arial", 12))
        other_button = tk.Button(
            frame_name, text="   Бошқа    ", font=("Arial", 12),
            command=lambda: entry_clicked(variable, entry))
        other_button.grid(row=label_row+offset, column=1, sticky="w", pady=5)
        return [variable, entry]

    return variable


def checkbox_question(frame_name, label_row, options, text, has_other=False):
    results = []
    label = tk.Label(frame_name, text=text)
    label.grid(row=label_row, column=0, columnspan=3, padx=10, pady=30)
    label.config(font=("Arial", 12))
    offset = 1
    for idx, option in enumerate(options):
        result = tk.IntVar()
        check_box = tk.Checkbutton(frame_name, text=option, variable=result, font=("Arial", 12))
        check_box.grid(row=label_row+offset, column=1, padx=10, pady=5, sticky="w")
        results.append(result)
        offset += 1

    # taking care of 'Other'
    if has_other:
        entry = tk.Entry(frame_name, width=100, validate="key", font=("Arial", 12))
        entry.configure(validatecommand=(entry.register(is_cyrillic),'%S','%d'))
        entry.grid(row=label_row+offset+1, column=0, columnspan=3, padx=10, pady=30)
        entry.configure(state=tk.DISABLED)
        button = tk.Button(
            frame_name, text="   Бошқа    ", font=("Arial", 12),
            command=lambda: entry.configure(state=tk.NORMAL))
        button.grid(row=label_row+offset, column=1, padx=10, pady=5, sticky="w")
        results.append(entry)
    return results


def inputting_questions(frame_name, label_row, num_options, text, has_difficult):
    results = []
    label = tk.Label(frame_name, text=text)
    label.grid(row=label_row, column=0, columnspan=3, padx=10, pady=30)
    label.config(font=("Arial", 12))
    for idx in range(1, num_options+1):
        entry = tk.Entry(frame_name, width=100, validate="key", font=("Arial", 12))
        entry.configure(validatecommand=(entry.register(is_cyrillic),'%S','%d'))
        entry.grid(row=label_row+idx, column=0, columnspan=3, padx=10, pady=10)
        results.append(entry)

    if has_difficult:
        is_difficult = tk.IntVar()
        check_box = tk.Checkbutton(frame_name, text="Жавоб беришга қийналаман", variable=is_difficult)
        check_box.configure(font=("Arial", 12))
        check_box.grid(row=label_row+num_options+1, column=1, padx=10, pady=5, sticky="w")
        results.append(is_difficult)
    return results
